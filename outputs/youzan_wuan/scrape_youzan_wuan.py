import csv
import json
import time
from pathlib import Path
from urllib.parse import urlencode
from urllib.request import Request, urlopen

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


OUT_DIR = Path(__file__).resolve().parent
KDT_ID = "113996920"
SHOP_HOST = "shop114189088.youzan.com"
UUID = "e7b48d32-4dab-4f3f-aecb-0fe66561b13b"
HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (iPhone; CPU iPhone OS 16_0 like Mac OS X) "
        "AppleWebKit/605.1.15 (KHTML, like Gecko) Version/16.0 Mobile/15E148 Safari/604.1"
    ),
    "Referer": f"https://shop114189088.m.youzan.com/wscshop/feature/goods/all?kdt_id={KDT_ID}",
}


def fetch_json(url):
    req = Request(url, headers=HEADERS)
    with urlopen(req, timeout=30) as resp:
        return json.loads(resp.read().decode("utf-8"))


def yuan(cents):
    if cents in ("", None):
        return None
    return round(float(cents) / 100, 2)


def compact_json(value):
    return json.dumps(value, ensure_ascii=False, separators=(",", ":"))


def get_shop():
    url = (
        f"https://{SHOP_HOST}/wscshopcore/extension/shop-info.json?"
        + urlencode({"kdt_id": KDT_ID, "referrer": "shop114189088.m.youzan.com", "kdtId": KDT_ID})
    )
    data = fetch_json(url)["data"]["shop"]
    return {
        "店铺id": data.get("kdtId"),
        "内部shop_id": data.get("shopId"),
        "名称": data.get("shopName", ""),
        "头像": data.get("logo", ""),
        "电话": data.get("phone") or data.get("mobile") or "",
        "地址": "".join([data.get("province", ""), data.get("city", ""), data.get("area", ""), data.get("address", "")]),
        "行业": data.get("businessName", ""),
        "行业id": data.get("business", ""),
        "原始店铺信息": compact_json(data),
    }


def all_goods_page(page):
    params = {
        "pageSize": 20,
        "page": page,
        "offlineId": 0,
        "openIndependentPrice": 0,
        "order": "",
        "json": 1,
        "uuid": UUID,
        "activityPriceIndependent": 1,
        "order_by": "algPVD30",
        "goodsType": 2,
        "isXhsLocalLife": "",
        "needActivity": 0,
        "clientSource": 2,
        "tagAlias": "",
        "needGroupFilter": "false",
        "needGoodsRank": "true",
        "supportCombo": "true",
        "excludedComboSubType": '["none"]',
        "kdt_id": KDT_ID,
    }
    url = f"https://{SHOP_HOST}/wscshop/showcase/goods/allGoods.json?" + urlencode(params)
    return fetch_json(url)


def detail(alias):
    params = {
        "app_id": "wxeb490c6f9dd01133",
        "kdt_id": KDT_ID,
        "alias": alias,
        "bizEnv": "retail",
        "platform": "weixin",
        "client": "weapp",
    }
    url = "https://retail-h5.youzan.com/wscgoods/tee-app/detail-v2.json?" + urlencode(params)
    try:
        return fetch_json(url).get("data") or {}
    except Exception as exc:
        return {"_detail_error": str(exc)}


def infer_category(title, shop_industry):
    accessory_words = ["蜡烛", "餐具", "发箍", "生日帽", "配件", "加拍", "专拍", "说明"]
    dessert_words = ["曲奇", "慕斯", "挞", "布丁", "脆脆"]
    if any(w in title for w in accessory_words):
        return "配件"
    if any(w in title for w in dessert_words):
        return "甜品"
    return shop_industry or "蛋糕烘焙"


def labels_from_good(good):
    labels = []
    rank = good.get("rankInfoDTO") or {}
    if rank.get("rankType") and rank.get("rankNo"):
        rank_name = {"hot_sale": "销量榜", "popular": "人气榜"}.get(rank.get("rankType"), rank.get("rankType"))
        labels.append(f"本店{rank_name}第{rank.get('rankNo')}")
    for info in good.get("activityInfos") or []:
        text = info.get("activityName") or info.get("name") or info.get("tag")
        if text and text not in labels:
            labels.append(text)
    if good.get("activityPrice") and good.get("price") and str(good.get("activityPrice")) != str(good.get("price")):
        labels.append("限时价")
    if "限时价" in str(good.get("messages", "")) and "限时价" not in labels:
        labels.append("限时价")
    return labels


def sku_options(sku_info):
    props = sku_info.get("props") or []
    skus = sku_info.get("skus") or []
    price_by_sku = {str(x.get("skuId")): yuan(x.get("price")) for x in sku_info.get("skuPrices") or []}
    groups = []
    combos = []
    value_lookup = {}
    for prop in props:
        key = prop.get("k_s")
        value_lookup[key] = {str(v.get("id")): v.get("name") for v in prop.get("v") or []}

    for sku in skus:
        combo = {"sku_id": sku.get("skuId"), "price": price_by_sku.get(str(sku.get("skuId")))}
        for prop in props:
            key = prop.get("k_s")
            raw = str(sku.get(key, "0"))
            if raw != "0":
                combo[prop.get("k")] = value_lookup.get(key, {}).get(raw, raw)
        combos.append(combo)

    for prop in props:
        key = prop.get("k_s")
        options = []
        for v in prop.get("v") or []:
            value_id = str(v.get("id"))
            prices = {
                price_by_sku.get(str(sku.get("skuId")))
                for sku in skus
                if str(sku.get(key, "0")) == value_id and price_by_sku.get(str(sku.get("skuId"))) is not None
            }
            item = {"name": v.get("name")}
            if len(prices) == 1:
                item["price"] = list(prices)[0]
            options.append(item)
        groups.append({"name": prop.get("k"), "options": options})
    return groups, combos


def normalize_good(good, shop):
    d = detail(good["alias"])
    gd = d.get("goodsData") or {}
    goods_base = gd.get("goods") or {}
    sku_info = gd.get("skuInfo") or {}
    groups, combos = sku_options(sku_info)
    prices = [v for v in [c.get("price") for c in combos] if v is not None]
    if not prices:
        prices = [float(good.get("activityPrice") or good.get("price") or 0)]
    images = [
        {"url": p.get("url"), "size": f"{p.get('width', '')}*{p.get('height', '')}".strip("*")}
        for p in (goods_base.get("pictures") or goods_base.get("picture") or good.get("picture") or [])
        if p.get("url")
    ]
    labels = labels_from_good(good)
    title = goods_base.get("title") or good.get("title", "")
    rank = good.get("rankInfoDTO") or {}
    return {
        "商品id": good.get("id"),
        "别名alias": good.get("alias"),
        "商品链接": good.get("url") or f"https://{SHOP_HOST}/v2/goods/{good.get('alias')}",
        "商品名称": title,
        "品类": infer_category(title, shop.get("行业", "")),
        "展示价格": float(good.get("activityPrice") or good.get("price") or min(prices)),
        "最低价": min(prices),
        "最高价": max(prices),
        "商品图": compact_json(images),
        "规格信息": compact_json(groups),
        "SKU明细": compact_json(combos),
        "热门": bool(rank),
        "推荐": bool(good.get("showRecommendedCard")),
        "标签": compact_json(labels),
        "销量": d.get("soldNum", good.get("totalSoldNum", 0)),
        "库存": good.get("totalStock", ""),
        "上架状态": good.get("soldStatus", ""),
        "详情抓取状态": "ok" if sku_info else d.get("_detail_error", "no_sku_info"),
    }


def write_outputs(shop, goods):
    json_path = OUT_DIR / "youzan_wuan_data.json"
    csv_path = OUT_DIR / "youzan_wuan_products.csv"
    xlsx_path = OUT_DIR / "youzan_wuan_data.xlsx"
    json_path.write_text(json.dumps({"shop": shop, "products": goods}, ensure_ascii=False, indent=2), encoding="utf-8")

    product_cols = list(goods[0].keys()) if goods else []
    with csv_path.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=product_cols)
        writer.writeheader()
        writer.writerows(goods)

    wb = Workbook()
    ws = wb.active
    ws.title = "店铺数据"
    shop_cols = list(shop.keys())
    ws.append(shop_cols)
    ws.append([shop[c] for c in shop_cols])

    ws2 = wb.create_sheet("商品数据")
    ws2.append(product_cols)
    for row in goods:
        ws2.append([row.get(c, "") for c in product_cols])

    for sheet in wb.worksheets:
        for cell in sheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="4F81BD")
        sheet.freeze_panes = "A2"
        for col_idx, column_cells in enumerate(sheet.columns, 1):
            width = min(max(len(str(cell.value or "")) for cell in column_cells) + 2, 60)
            sheet.column_dimensions[get_column_letter(col_idx)].width = width
        for row in sheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
    wb.save(xlsx_path)
    return json_path, csv_path, xlsx_path


def main():
    shop = get_shop()
    goods_by_id = {}
    page = 1
    while True:
        data = all_goods_page(page)
        payload = data.get("data") or {}
        items = payload.get("list") or []
        for item in items:
            goods_by_id[item["id"]] = item
        print(f"page={page} items={len(items)} has_more={payload.get('has_more')}")
        if not payload.get("has_more") or not items:
            break
        page += 1
        time.sleep(0.25)

    products = []
    for idx, good in enumerate(goods_by_id.values(), 1):
        print(f"detail {idx}/{len(goods_by_id)} {good.get('id')} {good.get('title')}")
        products.append(normalize_good(good, shop))
        time.sleep(0.12)

    paths = write_outputs(shop, products)
    print(json.dumps({"shop": shop["名称"], "products": len(products), "outputs": [str(p) for p in paths]}, ensure_ascii=False))


if __name__ == "__main__":
    main()
